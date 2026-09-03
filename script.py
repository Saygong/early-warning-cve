#!/usr/bin/env python3

# SIMPLE OpenCVE scraper.
# Reads vendor/product rows from assets_test.xlsx and generates a professional CVE report PDF.

import json
import re
import os
import requests
import urllib3
import time
import smtplib
from collections import deque
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from urllib.parse import quote_plus
from openpyxl import load_workbook
from dotenv import load_dotenv
from pdf_generator import write_pdf

load_dotenv()

# The site we query.
BASE_URL = 'https://app.opencve.io/api/v2'


LOOKBACK_DAYS = int(os.getenv('LOOKBACK_DAYS', '7'))
CVSS_THRESHOLD = float(os.getenv('CVSS_THRESHOLD'))
EPSS_THRESHOLD = float(os.getenv('EPSS_THRESHOLD'))
DIRECT_LINE_BASE_URL = os.getenv('DIRECT_LINE_BASE_URL')
DIRECT_LINE_TIMEOUT = int(os.getenv('DIRECT_LINE_TIMEOUT', '30'))
AGENT_RESPONSE_TIMEOUT = float(os.getenv('DIRECT_LINE_RESPONSE_TIMEOUT', '90'))
POLL_INTERVAL = 1.5
OPENCVE_CALL_LIMIT = 999
OPENCVE_CALL_WINDOW_SECONDS = 60 * 60
MAX_AI_VULNERABILITIES = 20

ASSETS_FILE_PATH = os.getenv('PATH_INPUT_FILE', 'assets_test.xlsx')
OUTPUT_PDF_FILE = os.getenv('PATH_OUTPUT_FILE', 'cves_last_week.pdf')

SMTP_HOST = os.getenv('SMTP_HOST')
SMTP_PORT = int(os.getenv('SMTP_PORT'))
#SMTP_USE_TLS = os.getenv('SMTP_USE_TLS', 'false').lower() in ('1', 'true', 'yes')
#SMTP_USERNAME = os.getenv('SMTP_USERNAME', '')
#SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', '')
EMAIL_SENDER = os.getenv('EMAIL_SENDER')
EMAIL_RECIPIENT = os.getenv('EMAIL_RECIPIENT')
EMAIL_SUBJECT = os.getenv('EMAIL_SUBJECT')
EMAIL_DISPLAY_NAME = os.getenv('EMAIL_DISPLAY_NAME')


# Use a single session for OpenCVE requests.
session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (compatible; early-warning-cve/1.0)',
    'Accept': 'application/json, text/plain, */*',
})

_token = os.getenv('OPENCVE_API_TOKEN')
if not _token:
    raise SystemExit('OPENCVE_API_TOKEN environment variable is required')
session.headers.update({'Authorization': f'Bearer {_token}'})

_opencve_call_times = deque()

_verify_env = os.getenv('OPENCVE_VERIFY')
VERIFY_SSL = not (_verify_env is None or _verify_env.lower() in ('0', 'false', 'no'))
session.verify = VERIFY_SSL
if not VERIFY_SSL:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def api_get(url, params=None):
    """GET helper for the OpenCVE API; returns parsed JSON."""
    now = time.monotonic()
    while _opencve_call_times and now - _opencve_call_times[0] >= OPENCVE_CALL_WINDOW_SECONDS:
        _opencve_call_times.popleft()

    if len(_opencve_call_times) >= OPENCVE_CALL_LIMIT:
        pause_seconds = max(
            OPENCVE_CALL_WINDOW_SECONDS,
            OPENCVE_CALL_WINDOW_SECONDS - (now - _opencve_call_times[0]),
        )
        print(f'[*] OpenCVE call limit reached; pausing for {pause_seconds / 60:.1f} minutes')
        time.sleep(max(0, pause_seconds))
        now = time.monotonic()
        while _opencve_call_times and now - _opencve_call_times[0] >= OPENCVE_CALL_WINDOW_SECONDS:
            _opencve_call_times.popleft()

    _opencve_call_times.append(now)
    resp = session.get(url, params=params, timeout=30)
    resp.raise_for_status()
    try:
        return resp.json()
    except ValueError:
        raise ValueError('Invalid JSON response from OpenCVE API')


def send_report_email(pdf_path, vuln_count):
    """Send the generated PDF to the configured email recipient."""
    if not EMAIL_RECIPIENT.strip():
        print('[!] Email not sent: EMAIL_RECIPIENT is not configured')
        return

    email_body = (f'''Good morning team,\n'''
                 f'''Please find attached the weekly CVE report covering the latest vulnerabilities identified across the monitored products.\n\n'''
                 f'''This report presents critical and high-severity CVEs discovered in the last {LOOKBACK_DAYS} days across monitored products. A total of {vuln_count} vulnerabilities were identified.\n\n'''

                 f'''Please review the attached report for further details on the identified vulnerabilities.\n\n'''

                 f'''Best regards,\n'''
                 f'''Cybersecurity Team''')


    message = EmailMessage()
    message['From'] = formataddr((EMAIL_DISPLAY_NAME, EMAIL_SENDER))
    message['To'] = EMAIL_RECIPIENT
    message['Subject'] = EMAIL_SUBJECT
    message.set_content(email_body)

    with open(pdf_path, 'rb') as attachment:
        message.add_attachment(
            attachment.read(),
            maintype='application',
            subtype='pdf',
            filename=Path(pdf_path).name,
        )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        #if SMTP_USE_TLS:
        #    smtp.starttls()
        #if SMTP_USERNAME:
        #    smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(message)



def slugify(text):
    """Convert a name to the form OpenCVE expects in query strings."""
    if not text:
        return ''

    text = text.strip().lower()
    text = re.sub(r'[^a-z0-9]+', '_', text)
    text = re.sub(r'_+', '_', text)
    return text.strip('_')


def read_assets_from_excel(path):
    """Read vendor and product names from the Excel file."""
    workbook = load_workbook(path)
    worksheet = workbook.active

    # Read the first row as headers.
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip().lower() if cell else '' for cell in header_row]

    # Only vendor_name and product_name are required.
    required = ['vendor_name', 'product_name']
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f'Missing Excel columns: {missing}')

    assets = []
    for row in worksheet.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            return assets

        vendor_name = str(row[headers.index('vendor_name')].value).strip()
        product_name = str(row[headers.index('product_name')].value).strip()

        assets.append({
            'vendor_name': vendor_name,
            'product_name': product_name,
        })

    return assets


def parse_date(text):
    """Try to parse a date string from the OpenCVE results."""
    if not text:
        return None

    text = text.strip()
    if not text or text.lower() in {'n/a', 'unknown', '-'}:
        return None

    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%b-%Y', '%b %d, %Y', '%d %b %Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def find_field(data, names):
    """Find the first matching field recursively in an API response."""
    if isinstance(data, dict):
        for name in names:
            if name in data and data[name] is not None:
                return data[name]
        for value in data.values():
            found = find_field(value, names)
            if found is not None:
                return found
    elif isinstance(data, list):
        for value in data:
            found = find_field(value, names)
            if found is not None:
                return found
    return None


def fetch_cves_from_api(vendor_slug, product_slug):
    """Fetch CVEs for a vendor/product via OpenCVE API (handles simple pagination).

    Returns a list of JSON CVE items (structure may vary slightly across API versions),
    so callers should access fields defensively.
    """
    url = f'{BASE_URL}/vendors/{quote_plus(vendor_slug)}/products/{quote_plus(product_slug)}/cves?page_size=5'
    items = []
    
    data = api_get(url)

    # Support several common pagination/response shapes
    page_items = []
    if isinstance(data, dict):
        if 'items' in data:
            page_items = data['items']
        elif 'data' in data:
            page_items = data['data']
        elif 'results' in data:
            page_items = data['results']
        else:
            # Some endpoints return a paginated object with 'items' under another key
            # or a single list under a named key. Try to heuristically pick the first list.
            for v in data.values():
                if isinstance(v, list):
                    page_items = v
                    break
    elif isinstance(data, list):
        page_items = data
    else:
        raise ValueError('Unexpected response shape from OpenCVE API')

    items.extend(page_items)

    return items


def fetch_cve_detail(cve_id):
    """Fetch the full CVE record, including CVSS and EPSS scores."""
    url = f'{BASE_URL}/cves/{quote_plus(cve_id)}'
    return api_get(url)


def filter_recent(rows):
    """Keep only rows created within the last LOOKBACK_DAYS."""
    cutoff = datetime.now(timezone.utc).date() - timedelta(days=LOOKBACK_DAYS)
    def created_date(r):
        ud = r.get('created_at')
        if not ud:
            return None
        return parse_date(str(ud))

    return [row for row in rows if created_date(row) and created_date(row) >= cutoff]


def filter_critical(rows):
    """Keep rows with CVSS >= 8.0 OR EPSS >= 0.1."""
    return [
        row for row in rows
        if (row.get('cvss') is not None and row['cvss'] >= CVSS_THRESHOLD)
        or (row.get('epss') is not None and row['epss'] >= EPSS_THRESHOLD)
    ]


def directline_request(method, url, token, **kwargs):
    """Execute a Direct Line request and return its JSON response."""
    headers = kwargs.pop('headers', {})
    headers['Authorization'] = f'Bearer {token}'
    response = requests.request(
        method,
        url,
        verify=VERIFY_SSL,
        headers=headers,
        timeout=DIRECT_LINE_TIMEOUT,
        **kwargs,
    )
    response.raise_for_status()
    return response.json() if response.content else {}


def business_impact_analysis_copilot(vulnerability):
    """Send one vulnerability to Copilot Studio and return its business impact analysis."""
    secret = os.getenv('COPILOT_AGENT_SECRET').strip()
    if not secret:
        raise RuntimeError('COPILOT_AGENT_SECRET environment variable is required')

    token_data = directline_request(
        'POST',
        f'{DIRECT_LINE_BASE_URL}/tokens/generate',
        secret,
        headers={'Content-Type': 'application/json'},
    )
    token = token_data.get('token')
    if not token:
        raise RuntimeError('Direct Line token missing from response')

    conversation_data = directline_request(
        'POST',
        f'{DIRECT_LINE_BASE_URL}/conversations',
        token,
    )
    conversation_id = conversation_data.get('conversationId')
    if not conversation_id:
        raise RuntimeError('Direct Line conversationId missing from response')

    prompt = (
        'Analizza la seguente vulnerabilita e produci un testo in linguaggio business '
        'di massimo 70 parole in cui descrivi i potenziali impatti, se esiste un '
        'exploit pubblico e cosa succede se non viene patchata. Se un dato manca nel '
        'JSON, dichiaralo chiaramente.\n\nJSON:\n'
        + json.dumps(vulnerability, ensure_ascii=False, separators=(',', ':'))
    )
    activity_data = directline_request(
        'POST',
        f'{DIRECT_LINE_BASE_URL}/conversations/{conversation_id}/activities',
        token,
        headers={'Content-Type': 'application/json'},
        json={
            'type': 'message',
            'locale': 'it-IT',
            'from': {'id': 'script', 'role': 'user'},
            'text': prompt,
        },
    )
    sent_activity_id = activity_data.get('id')
    if not sent_activity_id:
        raise RuntimeError('Direct Line activity id missing from response')

    deadline = time.monotonic() + AGENT_RESPONSE_TIMEOUT
    watermark = None
    while time.monotonic() < deadline:
        params = {'watermark': watermark} if watermark else None
        activities_data = directline_request(
            'GET',
            f'{DIRECT_LINE_BASE_URL}/conversations/{conversation_id}/activities',
            token,
            params=params,
        )
        watermark = activities_data.get('watermark')
        for activity in activities_data.get('activities', []):
            if activity.get('id') == sent_activity_id:
                continue
            sender = activity.get('from') or {}
            if activity.get('type') == 'message' and activity.get('text') and sender.get('id') != 'script':
                return activity['text'].strip()
        time.sleep(POLL_INTERVAL)

    raise TimeoutError(f'No Copilot response within {AGENT_RESPONSE_TIMEOUT} seconds')

def remediation_suggestions_copilot(vulnerability):
    """Send one vulnerability to Copilot Studio and return its remediation suggestions."""
    secret = os.getenv('COPILOT_AGENT_SECRET').strip()
    if not secret:
        raise RuntimeError('COPILOT_AGENT_SECRET environment variable is required')

    token_data = directline_request(
        'POST',
        f'{DIRECT_LINE_BASE_URL}/tokens/generate',
        secret,
        headers={'Content-Type': 'application/json'},
    )
    token = token_data.get('token')
    if not token:
        raise RuntimeError('Direct Line token missing from response')

    conversation_data = directline_request(
        'POST',
        f'{DIRECT_LINE_BASE_URL}/conversations',
        token,
    )
    conversation_id = conversation_data.get('conversationId')
    if not conversation_id:
        raise RuntimeError('Direct Line conversationId missing from response')

    prompt = (
        'Senza riportare la descrizione e le informazioni generali della vulnerabilità, produci un testo in linguaggio business '
        'di massimo 50 parole in cui descrivi e consigli le attività di remediation o, se non possibile, le misure di mitigazione.\n\nJSON:\n'
        + json.dumps(vulnerability, ensure_ascii=False, separators=(',', ':'))
    )
    activity_data = directline_request(
        'POST',
        f'{DIRECT_LINE_BASE_URL}/conversations/{conversation_id}/activities',
        token,
        headers={'Content-Type': 'application/json'},
        json={
            'type': 'message',
            'locale': 'it-IT',
            'from': {'id': 'script', 'role': 'user'},
            'text': prompt,
        },
    )
    sent_activity_id = activity_data.get('id')
    if not sent_activity_id:
        raise RuntimeError('Direct Line activity id missing from response')

    deadline = time.monotonic() + AGENT_RESPONSE_TIMEOUT
    watermark = None
    while time.monotonic() < deadline:
        params = {'watermark': watermark} if watermark else None
        activities_data = directline_request(
            'GET',
            f'{DIRECT_LINE_BASE_URL}/conversations/{conversation_id}/activities',
            token,
            params=params,
        )
        watermark = activities_data.get('watermark')
        for activity in activities_data.get('activities', []):
            if activity.get('id') == sent_activity_id:
                continue
            sender = activity.get('from') or {}
            if activity.get('type') == 'message' and activity.get('text') and sender.get('id') != 'script':
                return activity['text'].strip()
        time.sleep(POLL_INTERVAL)

    raise TimeoutError(f'No Copilot response within {AGENT_RESPONSE_TIMEOUT} seconds')

def extract_cve_fields(item, score):
    """Normalize a CVE item from the API into our expected dict keys."""
    def get_first(keys, default=''):
        for k in keys:
            if k in item and item[k] is not None:
                return item[k]
        return default

    cve_id = get_first(['cve_id'])
    created_at = get_first(['created_at'])
    description = get_first(['description'])
    

    if not score:

        return {
            'cve_id': str(cve_id),
            'created_at': str(created_at),
            'cvss': None,
            'epss': None,
            'description': str(description),
        }
    else:

        return {
            'cve_id': str(cve_id),
            'created_at': str(created_at),
            'cvss': item.get('metrics', {}).get('cvssV3_1', {}).get('data').get('score'),
            'epss': item.get('metrics', {}).get('epss', {}).get('data').get('score'),
            'description': str(description),
        }
    


def get_cves_for_asset(asset):
    """Fetch CVE candidates for a single asset using the OpenCVE API."""
    vendor_name = asset['vendor_name']
    product_name = asset['product_name']
    vendor_slug = slugify(vendor_name)
    product_slug = slugify(product_name)

    print(f'[*] Querying OpenCVE API: vendor={vendor_name}, product={product_name}')
    items = fetch_cves_from_api(vendor_slug, product_slug)

    # The list endpoint has publication dates; fetch details only for recent CVEs.
    rows_with_json = []
    for item in items:
        row = extract_cve_fields(item, score=False)
        row['_raw_json'] = item
        rows_with_json.append(row)

    rows = filter_recent(rows_with_json)
    detailed_rows = []
    for row in rows:
        detail = fetch_cve_detail(row['cve_id'])
        vulnerability_json = {**row['_raw_json'], **detail}
        detailed_row = extract_cve_fields(vulnerability_json, score=True)
        detailed_row['_raw_json'] = vulnerability_json
        detailed_rows.append(detailed_row)

    rows = filter_critical(detailed_rows)

    output = []
    for row in rows:
        output.append({
            'vendor_name': vendor_name,
            'product_name': product_name,
            'cve_id': row.get('cve_id', ''),
            'created_at': row.get('created_at', ''),
            'cvss': row.get('cvss'),
            'epss': row.get('epss'),
            'description': row.get('description', ''),
            '_raw_json': row['_raw_json'],
        })

    print(f'[*] Found {len(output)} CVEs for {vendor_name} / {product_name}')
    return output


def build_report_rows(candidates):
    """Sort candidates and enrich at most the top 20 with Copilot."""
    def score(value):
        try:
            return float(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    candidates.sort(
        key=lambda row: (
            score(row.get('cvss')) is not None,
            score(row.get('cvss')) or 0,
            score(row.get('epss')) is not None,
            score(row.get('epss')) or 0,
        ),
        reverse=True,
    )

    ai_enabled = os.getenv('USE_COPILOT', '').strip().lower() in ('1', 'true', 'yes')
    if ai_enabled and not os.getenv('COPILOT_AGENT_SECRET', '').strip():
        raise SystemExit('COPILOT_AGENT_SECRET environment variable is required')

    report_rows = []
    for index, row in enumerate(candidates):
        report_row = [
            row['vendor_name'],
            row['product_name'],
            row['cve_id'],
            row['created_at'],
            '' if row['cvss'] is None else str(row['cvss']),
            '' if row['epss'] is None else str(row['epss']),
            row['description'],
        ]

        if ai_enabled:
            if index < MAX_AI_VULNERABILITIES:
                report_row.extend([
                    business_impact_analysis_copilot(row['_raw_json']),
                    remediation_suggestions_copilot(row['_raw_json']),
                ])
            else:
                report_row.extend(['N/A', 'N/A'])

        report_rows.append(report_row)

    return report_rows


def main():
    """Main script flow: read assets, query OpenCVE, and generate PDF report."""
    assets_path = Path(ASSETS_FILE_PATH)
    if not assets_path.exists():
        raise SystemExit(f'Excel file not found: {assets_path}')

    # Return a list of dicts with 'vendor_name' and 'product_name' keys. 
    assets = read_assets_from_excel(assets_path)
    if not assets:
        raise SystemExit('No valid vendor/product rows found in the Excel file')

    candidates = []
    for asset in assets:
        try:
            candidates.extend(get_cves_for_asset(asset))
        except Exception as exc:
            print(f'[!] Skipping {asset["vendor_name"]} / {asset["product_name"]}: {exc}')

    all_rows = build_report_rows(candidates)
    output_file_name = OUTPUT_PDF_FILE + f'_{datetime.now().strftime("%d-%m-%Y")}.pdf'
    write_pdf(output_file_name, all_rows)
    print(f'[*] Generated report with {len(all_rows)} CVEs')
    if all_rows:
        send_report_email(output_file_name, len(all_rows))
        print(f'[*] Report emailed to selected recipients')
    else:
        print('[*] Report email skipped: no vulnerabilities found')


if __name__ == '__main__':
    main()
